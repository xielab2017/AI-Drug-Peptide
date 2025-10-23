#!/usr/bin/env python3
"""
AI-Drug Peptide - 报告生成模块
包含报告生成器、可视化引擎、图表生成器和导出管理器
"""

import os
import json
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass, asdict
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages

# Excel支持
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Plotly支持
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    import plotly.offline as pyo
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

logger = logging.getLogger(__name__)

@dataclass
class ReportData:
    """报告数据"""
    protein_id: str
    analysis_results: Dict[str, Any]
    metadata: Dict[str, Any]
    generated_at: str

@dataclass
class ChartConfig:
    """图表配置"""
    chart_type: str
    title: str
    x_label: str
    y_label: str
    data: Dict[str, Any]
    style: Dict[str, Any] = None

class ReportGenerator:
    """报告生成器"""
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.output_dir = Path(self.config.get('output_dir', './data/output'))
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # 设置matplotlib样式
        plt.style.use('default')
        sns.set_palette("husl")
    
    async def generate_report(self, request: Any, results: Dict[str, Any]) -> Dict[str, str]:
        """生成分析报告"""
        logger.info(f"Generating report for protein: {request.protein_id}")
        
        report_files = {}
        
        try:
            # 准备报告数据
            report_data = ReportData(
                protein_id=request.protein_id,
                analysis_results=results,
                metadata={
                    'analysis_steps': request.analysis_steps,
                    'species_id': request.species_id,
                    'confidence_threshold': request.confidence_threshold,
                    'energy_threshold': request.energy_threshold,
                    'conservation_threshold': request.conservation_threshold,
                    'target_species': request.target_species
                },
                generated_at=datetime.now().isoformat()
            )
            
            # 生成JSON报告
            json_file = self._generate_json_report(report_data)
            report_files['json'] = str(json_file)
            
            # 生成Excel报告
            if EXCEL_AVAILABLE:
                excel_file = self._generate_excel_report(report_data)
                report_files['excel'] = str(excel_file)
            
            # 生成PDF报告
            pdf_file = self._generate_pdf_report(report_data)
            report_files['pdf'] = str(pdf_file)
            
            # 生成HTML报告
            if PLOTLY_AVAILABLE:
                html_file = self._generate_html_report(report_data)
                report_files['html'] = str(html_file)
            
            logger.info(f"Report generation completed: {len(report_files)} files created")
            return report_files
            
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            raise
    
    def _generate_json_report(self, report_data: ReportData) -> Path:
        """生成JSON报告"""
        filename = f"analysis_report_{report_data.protein_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path = self.output_dir / filename
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(asdict(report_data), f, default=str, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON report saved: {file_path}")
        return file_path
    
    def _generate_excel_report(self, report_data: ReportData) -> Path:
        """生成Excel报告"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available for Excel report generation")
        
        filename = f"analysis_report_{report_data.protein_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = self.output_dir / filename
        
        # 创建工作簿
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # 删除默认工作表
        
        # 创建概览表
        self._create_summary_sheet(workbook, report_data)
        
        # 创建详细结果表
        self._create_results_sheet(workbook, report_data)
        
        # 创建图表表
        self._create_charts_sheet(workbook, report_data)
        
        # 保存工作簿
        workbook.save(file_path)
        
        logger.info(f"Excel report saved: {file_path}")
        return file_path
    
    def _create_summary_sheet(self, workbook, report_data: ReportData):
        """创建概览表"""
        summary_sheet = workbook.create_sheet("Summary")
        
        # 标题
        summary_data = [
            ['AI-Drug Peptide 分析报告'],
            [''],
            ['蛋白质ID:', report_data.protein_id],
            ['分析时间:', report_data.generated_at],
            ['分析步骤:', ', '.join(report_data.metadata['analysis_steps'])],
            ['物种ID:', report_data.metadata['species_id']],
            ['置信度阈值:', report_data.metadata['confidence_threshold']],
            ['能量阈值:', report_data.metadata['energy_threshold']],
            ['保守性阈值:', report_data.metadata['conservation_threshold']],
            ['目标物种:', ', '.join(report_data.metadata['target_species'])],
            [''],
            ['分析结果概览:']
        ]
        
        # 添加结果统计
        results = report_data.analysis_results
        if 'string_analysis' in results:
            string_result = results['string_analysis']
            summary_data.extend([
                ['STRING相互作用数量:', string_result.get('total_interactions', 0)],
                ['平均置信度:', f"{np.mean(string_result.get('confidence_scores', [0])):.3f}"]
            ])
        
        if 'docking_analysis' in results:
            docking_results = results['docking_analysis']
            if isinstance(docking_results, list) and docking_results:
                avg_energy = np.mean([r.get('binding_energy', 0) for r in docking_results])
                high_affinity_count = sum(1 for r in docking_results if r.get('high_affinity', False))
                summary_data.extend([
                    ['平均结合能:', f"{avg_energy:.2f} kcal/mol"],
                    ['高亲和力受体数量:', high_affinity_count]
                ])
        
        if 'conservation_analysis' in results:
            conservation_result = results['conservation_analysis']
            summary_data.extend([
                ['平均保守性:', f"{conservation_result.get('avg_conservation', 0):.3f}"],
                ['保守性状态:', '保守' if conservation_result.get('is_conservative', False) else '非保守']
            ])
        
        # 写入数据
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                self._style_summary_cell(cell, row_num)
        
        # 自动调整列宽
        for column in summary_sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            column_letter = column[0].column_letter
            summary_sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_results_sheet(self, workbook, report_data: ReportData):
        """创建详细结果表"""
        results_sheet = workbook.create_sheet("Detailed Results")
        
        # 准备数据
        all_data = []
        results = report_data.analysis_results
        
        # STRING结果
        if 'string_analysis' in results:
            string_result = results['string_analysis']
            for i, protein in enumerate(string_result.get('interacting_proteins', [])):
                all_data.append({
                    'Analysis_Type': 'STRING',
                    'Receptor_ID': protein.get('receptor_id', ''),
                    'Gene_Name': protein.get('gene_name', ''),
                    'Confidence': protein.get('confidence', 0),
                    'Literature_Support': string_result.get('literature_support', [0])[i] if i < len(string_result.get('literature_support', [])) else 0,
                    'UniProt_ID': protein.get('uniprot_id', ''),
                    'PDB_ID': protein.get('pdb_id', '')
                })
        
        # 对接结果
        if 'docking_analysis' in results:
            docking_results = results['docking_analysis']
            if isinstance(docking_results, list):
                for result in docking_results:
                    all_data.append({
                        'Analysis_Type': 'Docking',
                        'Receptor_ID': result.get('receptor_id', ''),
                        'Gene_Name': '',
                        'Confidence': '',
                        'Literature_Support': '',
                        'Binding_Energy': result.get('binding_energy', 0),
                        'Success_Rate': result.get('success_rate', 0),
                        'Conformations': result.get('conformations', 0),
                        'High_Affinity': result.get('high_affinity', False)
                    })
        
        # 保守性结果
        if 'conservation_analysis' in results:
            conservation_result = results['conservation_analysis']
            all_data.append({
                'Analysis_Type': 'Conservation',
                'Receptor_ID': '',
                'Gene_Name': '',
                'Confidence': '',
                'Literature_Support': '',
                'Species_Count': conservation_result.get('species_count', 0),
                'Avg_Conservation': conservation_result.get('avg_conservation', 0),
                'Is_Conservative': conservation_result.get('is_conservative', False)
            })
        
        # 创建DataFrame
        if all_data:
            df = pd.DataFrame(all_data)
            
            # 写入Excel
            for row in dataframe_to_rows(df, index=False, header=True):
                results_sheet.append(row)
            
            # 样式化
            self._style_results_sheet(results_sheet, len(df.columns))
    
    def _create_charts_sheet(self, workbook, report_data: ReportData):
        """创建图表表"""
        charts_sheet = workbook.create_sheet("Charts")
        
        # 添加图表说明
        charts_data = [
            ['图表说明'],
            [''],
            ['本工作表包含分析过程中生成的各种图表和可视化结果。'],
            [''],
            ['图表类型:'],
            ['1. STRING相互作用网络图'],
            ['2. 分子对接结合能分布图'],
            ['3. 保守性分析热图'],
            ['4. 综合评分雷达图'],
            [''],
            ['注意: 详细的交互式图表请查看HTML报告文件。']
        ]
        
        for row_num, row_data in enumerate(charts_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = charts_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # 标题
                    cell.font = Font(bold=True, size=14)
    
    def _generate_pdf_report(self, report_data: ReportData) -> Path:
        """生成PDF报告"""
        filename = f"analysis_report_{report_data.protein_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = self.output_dir / filename
        
        with PdfPages(file_path) as pdf:
            # 创建图表
            self._create_summary_chart(pdf, report_data)
            self._create_interaction_chart(pdf, report_data)
            self._create_docking_chart(pdf, report_data)
            self._create_conservation_chart(pdf, report_data)
        
        logger.info(f"PDF report saved: {file_path}")
        return file_path
    
    def _create_summary_chart(self, pdf, report_data: ReportData):
        """创建概览图表"""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # 分析步骤统计
        steps = report_data.metadata['analysis_steps']
        step_names = {
            'string': 'STRING分析',
            'docking': '分子对接',
            'conservation': '保守性分析',
            'secretion': '分泌分析'
        }
        
        step_labels = [step_names.get(step, step) for step in steps]
        step_counts = [1] * len(steps)  # 每个步骤都完成了
        
        bars = ax.bar(step_labels, step_counts, color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728'])
        ax.set_title(f'蛋白质 {report_data.protein_id} 分析步骤完成情况', fontsize=14, fontweight='bold')
        ax.set_ylabel('完成状态')
        ax.set_ylim(0, 1.2)
        
        # 添加数值标签
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.05,
                   '✓ 完成', ha='center', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches='tight')
        plt.close()
    
    def _create_interaction_chart(self, pdf, report_data: ReportData):
        """创建相互作用图表"""
        if 'string_analysis' not in report_data.analysis_results:
            return
        
        string_result = report_data.analysis_results['string_analysis']
        confidence_scores = string_result.get('confidence_scores', [])
        
        if not confidence_scores:
            return
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # 置信度分布直方图
        ax.hist(confidence_scores, bins=10, alpha=0.7, color='skyblue', edgecolor='black')
        ax.axvline(np.mean(confidence_scores), color='red', linestyle='--', 
                  label=f'平均值: {np.mean(confidence_scores):.3f}')
        ax.set_xlabel('置信度')
        ax.set_ylabel('相互作用数量')
        ax.set_title(f'蛋白质 {report_data.protein_id} STRING相互作用置信度分布')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches='tight')
        plt.close()
    
    def _create_docking_chart(self, pdf, report_data: ReportData):
        """创建对接图表"""
        if 'docking_analysis' not in report_data.analysis_results:
            return
        
        docking_results = report_data.analysis_results['docking_analysis']
        if not isinstance(docking_results, list) or not docking_results:
            return
        
        binding_energies = [r.get('binding_energy', 0) for r in docking_results]
        receptor_ids = [r.get('receptor_id', f'Receptor_{i}') for i, r in enumerate(docking_results)]
        
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # 结合能柱状图
        bars = ax.bar(range(len(receptor_ids)), binding_energies, 
                     color=['red' if e < -7 else 'orange' if e < -5 else 'green' for e in binding_energies])
        
        ax.set_xlabel('受体')
        ax.set_ylabel('结合能 (kcal/mol)')
        ax.set_title(f'蛋白质 {report_data.protein_id} 分子对接结合能')
        ax.set_xticks(range(len(receptor_ids)))
        ax.set_xticklabels(receptor_ids, rotation=45, ha='right')
        
        # 添加阈值线
        ax.axhline(y=-7, color='red', linestyle='--', alpha=0.7, label='高亲和力阈值')
        
        # 添加数值标签
        for i, (bar, energy) in enumerate(zip(bars, binding_energies)):
            ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.1,
                   f'{energy:.2f}', ha='center', va='bottom')
        
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches='tight')
        plt.close()
    
    def _create_conservation_chart(self, pdf, report_data: ReportData):
        """创建保守性图表"""
        if 'conservation_analysis' not in report_data.analysis_results:
            return
        
        conservation_result = report_data.analysis_results['conservation_analysis']
        conservation_scores = conservation_result.get('conservation_scores', [])
        
        if not conservation_scores:
            return
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # 保守性得分分布
        species_names = conservation_result.get('species_names', [])
        if len(species_names) >= 2:
            # 创建物种对比图
            x_pos = range(len(conservation_scores))
            bars = ax.bar(x_pos, conservation_scores, alpha=0.7, color='lightgreen', edgecolor='black')
            
            ax.set_xlabel('物种对')
            ax.set_ylabel('保守性得分')
            ax.set_title(f'蛋白质 {report_data.protein_id} 保守性分析')
            ax.set_xticks(x_pos)
            ax.set_xticklabels([f'Species {i+1}' for i in range(len(conservation_scores))])
            
            # 添加平均值线
            avg_conservation = conservation_result.get('avg_conservation', 0)
            ax.axhline(y=avg_conservation, color='red', linestyle='--', 
                      label=f'平均值: {avg_conservation:.3f}')
            
            # 添加阈值线
            threshold = report_data.metadata.get('conservation_threshold', 0.8)
            ax.axhline(y=threshold, color='orange', linestyle=':', 
                      label=f'阈值: {threshold}')
            
            # 添加数值标签
            for bar, score in zip(bars, conservation_scores):
                ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.01,
                       f'{score:.3f}', ha='center', va='bottom')
            
            ax.legend()
            ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches='tight')
        plt.close()
    
    def _generate_html_report(self, report_data: ReportData) -> Path:
        """生成HTML报告"""
        if not PLOTLY_AVAILABLE:
            raise ImportError("plotly not available for HTML report generation")
        
        filename = f"analysis_report_{report_data.protein_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        file_path = self.output_dir / filename
        
        # 创建交互式图表
        fig = make_subplots(
            rows=2, cols=2,
            subplot_titles=('STRING相互作用', '分子对接', '保守性分析', '综合评分'),
            specs=[[{"type": "bar"}, {"type": "bar"}],
                   [{"type": "bar"}, {"type": "scatter"}]]
        )
        
        # 添加图表内容
        self._add_interactive_charts(fig, report_data)
        
        # 更新布局
        fig.update_layout(
            title=f'蛋白质 {report_data.protein_id} 分析报告',
            showlegend=True,
            height=800
        )
        
        # 保存HTML文件
        pyo.plot(fig, filename=str(file_path), auto_open=False)
        
        logger.info(f"HTML report saved: {file_path}")
        return file_path
    
    def _add_interactive_charts(self, fig, report_data: ReportData):
        """添加交互式图表"""
        results = report_data.analysis_results
        
        # STRING相互作用图
        if 'string_analysis' in results:
            string_result = results['string_analysis']
            confidence_scores = string_result.get('confidence_scores', [])
            if confidence_scores:
                fig.add_trace(
                    go.Bar(x=list(range(len(confidence_scores))), y=confidence_scores,
                          name='置信度', marker_color='lightblue'),
                    row=1, col=1
                )
        
        # 分子对接图
        if 'docking_analysis' in results:
            docking_results = results['docking_analysis']
            if isinstance(docking_results, list) and docking_results:
                receptor_ids = [r.get('receptor_id', f'R{i}') for i, r in enumerate(docking_results)]
                binding_energies = [r.get('binding_energy', 0) for r in docking_results]
                
                fig.add_trace(
                    go.Bar(x=receptor_ids, y=binding_energies,
                          name='结合能', marker_color='lightcoral'),
                    row=1, col=2
                )
        
        # 保守性分析图
        if 'conservation_analysis' in results:
            conservation_result = results['conservation_analysis']
            conservation_scores = conservation_result.get('conservation_scores', [])
            if conservation_scores:
                fig.add_trace(
                    go.Bar(x=[f'Pair {i+1}' for i in range(len(conservation_scores))], 
                          y=conservation_scores,
                          name='保守性', marker_color='lightgreen'),
                    row=2, col=1
                )
        
        # 综合评分雷达图
        scores = []
        labels = []
        
        if 'string_analysis' in results:
            string_result = results['string_analysis']
            avg_confidence = np.mean(string_result.get('confidence_scores', [0]))
            scores.append(avg_confidence)
            labels.append('STRING置信度')
        
        if 'docking_analysis' in results:
            docking_results = results['docking_analysis']
            if isinstance(docking_results, list) and docking_results:
                avg_energy = np.mean([r.get('binding_energy', 0) for r in docking_results])
                # 转换为0-1评分（-15到-1映射到0-1）
                energy_score = max(0, min(1, (avg_energy + 15) / 14))
                scores.append(energy_score)
                labels.append('结合能评分')
        
        if 'conservation_analysis' in results:
            conservation_result = results['conservation_analysis']
            conservation_score = conservation_result.get('avg_conservation', 0)
            scores.append(conservation_score)
            labels.append('保守性评分')
        
        if scores and labels:
            fig.add_trace(
                go.Scatterpolar(r=scores, theta=labels, fill='toself', name='综合评分'),
                row=2, col=2
            )
    
    def _style_summary_cell(self, cell, row_num):
        """样式化概览表单元格"""
        if row_num == 1:  # 标题
            cell.font = Font(bold=True, size=16)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, size=16, color="FFFFFF")
        elif ':' in str(cell.value or '') and not str(cell.value or '').endswith(':'):  # 标签
            cell.font = Font(bold=True)
        
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_results_sheet(self, sheet, num_cols):
        """样式化结果表"""
        # 标题样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 样式化标题行
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 样式化数据行
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'

class VisualizationEngine:
    """可视化引擎"""
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.output_dir = Path(self.config.get('output_dir', './data/output'))
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def create_network_graph(self, interactions: List[Dict[str, Any]]) -> Optional[Path]:
        """创建网络图"""
        try:
            import networkx as nx
            
            # 创建网络图
            G = nx.Graph()
            
            # 添加节点和边
            for interaction in interactions:
                protein_a = interaction.get('protein_id_a', '')
                protein_b = interaction.get('protein_id_b', '')
                confidence = interaction.get('confidence', 0)
                
                if protein_a and protein_b:
                    G.add_edge(protein_a, protein_b, weight=confidence)
            
            # 绘制网络图
            plt.figure(figsize=(12, 8))
            pos = nx.spring_layout(G, k=1, iterations=50)
            
            # 绘制边
            edges = G.edges()
            weights = [G[u][v]['weight'] for u, v in edges]
            nx.draw_networkx_edges(G, pos, width=weights, alpha=0.6, edge_color='gray')
            
            # 绘制节点
            nx.draw_networkx_nodes(G, pos, node_color='lightblue', 
                                  node_size=1000, alpha=0.8)
            
            # 绘制标签
            nx.draw_networkx_labels(G, pos, font_size=8, font_weight='bold')
            
            plt.title('蛋白质相互作用网络图', fontsize=16, fontweight='bold')
            plt.axis('off')
            
            # 保存图片
            filename = f"network_graph_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            file_path = self.output_dir / filename
            plt.savefig(file_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            logger.info(f"Network graph saved: {file_path}")
            return file_path
            
        except ImportError:
            logger.warning("NetworkX not available for network graph creation")
            return None
        except Exception as e:
            logger.error(f"Failed to create network graph: {e}")
            return None
    
    def create_heatmap(self, data: pd.DataFrame, title: str = "Heatmap") -> Optional[Path]:
        """创建热图"""
        try:
            plt.figure(figsize=(10, 8))
            sns.heatmap(data, annot=True, cmap='viridis', fmt='.3f')
            plt.title(title, fontsize=14, fontweight='bold')
            plt.tight_layout()
            
            # 保存图片
            filename = f"heatmap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            file_path = self.output_dir / filename
            plt.savefig(file_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            logger.info(f"Heatmap saved: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Failed to create heatmap: {e}")
            return None

class ExportManager:
    """导出管理器"""
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.output_dir = Path(self.config.get('output_dir', './data/output'))
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_csv(self, data: Union[Dict, List, pd.DataFrame], filename: str) -> Path:
        """导出为CSV格式"""
        file_path = self.output_dir / f"{filename}.csv"
        
        if isinstance(data, pd.DataFrame):
            data.to_csv(file_path, index=False, encoding='utf-8')
        elif isinstance(data, list):
            df = pd.DataFrame(data)
            df.to_csv(file_path, index=False, encoding='utf-8')
        elif isinstance(data, dict):
            # 展平字典数据
            flattened_data = self._flatten_dict(data)
            df = pd.DataFrame([flattened_data])
            df.to_csv(file_path, index=False, encoding='utf-8')
        
        logger.info(f"Data exported to CSV: {file_path}")
        return file_path
    
    def export_to_json(self, data: Any, filename: str) -> Path:
        """导出为JSON格式"""
        file_path = self.output_dir / f"{filename}.json"
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, default=str, indent=2, ensure_ascii=False)
        
        logger.info(f"Data exported to JSON: {file_path}")
        return file_path
    
    def _flatten_dict(self, data: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
        """展平嵌套字典"""
        items = []
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)

# 报告生成工厂类
class ReportGeneratorFactory:
    """报告生成器工厂"""
    
    @staticmethod
    def create_generator(config: Dict[str, Any] = None) -> ReportGenerator:
        """创建报告生成器"""
        return ReportGenerator(config)
    
    @staticmethod
    def create_visualization_engine(config: Dict[str, Any] = None) -> VisualizationEngine:
        """创建可视化引擎"""
        return VisualizationEngine(config)
    
    @staticmethod
    def create_export_manager(config: Dict[str, Any] = None) -> ExportManager:
        """创建导出管理器"""
        return ExportManager(config)

# 使用示例
if __name__ == "__main__":
    # 配置示例
    config = {
        'output_dir': './data/output',
        'log_level': 'INFO'
    }
    
    # 创建报告生成器
    generator = ReportGeneratorFactory.create_generator(config)
    
    # 模拟分析请求和结果
    class MockRequest:
        def __init__(self):
            self.protein_id = 'THBS4'
            self.analysis_steps = ['string', 'docking', 'conservation']
            self.species_id = 9606
            self.confidence_threshold = 0.9
            self.energy_threshold = -7.0
            self.conservation_threshold = 0.8
            self.target_species = ['human', 'mouse']
    
    mock_request = MockRequest()
    mock_results = {
        'string_analysis': {
            'total_interactions': 5,
            'confidence_scores': [0.95, 0.88, 0.82, 0.79, 0.76],
            'interacting_proteins': [
                {'receptor_id': 'EGFR', 'gene_name': 'EGFR', 'confidence': 0.95},
                {'receptor_id': 'MET', 'gene_name': 'MET', 'confidence': 0.88},
                {'receptor_id': 'KDR', 'gene_name': 'KDR', 'confidence': 0.82}
            ]
        },
        'docking_analysis': [
            {'receptor_id': 'EGFR', 'binding_energy': -8.5, 'high_affinity': True},
            {'receptor_id': 'MET', 'binding_energy': -7.2, 'high_affinity': True},
            {'receptor_id': 'KDR', 'binding_energy': -6.8, 'high_affinity': False}
        ],
        'conservation_analysis': {
            'avg_conservation': 0.85,
            'is_conservative': True,
            'conservation_scores': [0.82, 0.88, 0.85]
        }
    }
    
    # 生成报告
    async def generate_demo_report():
        try:
            report_files = await generator.generate_report(mock_request, mock_results)
            print("✓ Report generation completed successfully")
            print("Generated files:")
            for format_type, file_path in report_files.items():
                print(f"  {format_type.upper()}: {file_path}")
        except Exception as e:
            print(f"✗ Report generation failed: {e}")
    
    # 运行示例
    import asyncio
    asyncio.run(generate_demo_report())
