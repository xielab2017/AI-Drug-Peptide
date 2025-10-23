#!/usr/bin/env python3
"""
Step 4: Results Merging and Receptor Priority Ranking
功能：合并步骤1-3的结果，生成受体优先级列表并更新知识图谱

核心功能：
1. 数据合并：通过"受体ID"关联三表，保留同时存在于三个步骤的受体
2. 优先级评分：按公式计算综合得分（结合能权重40% + 保守性权重30% + 文献支持度权重30%）
3. 排序输出：按综合得分降序排列，取Top3受体标注为"核心受体"
4. 更新Neo4j：用py2neo创建关系"(蛋白)-[结合能{值:xx}]->(受体)""(受体)-[保守性{值:xx}]->(同源受体)"

作者：AI-Drug Peptide Project
日期：2024
"""

import os
import sys
import json
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass
from datetime import datetime
import warnings

# Neo4j integration (optional)
try:
    from py2neo import Graph, Node, Relationship, Subgraph
    from neo4j import GraphDatabase
    NEO4J_AVAILABLE = True
except ImportError:
    NEO4J_AVAILABLE = False
    print("Warning: Neo4j integration not available. Install py2neo for Neo4j support.")

# Excel reporting
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('./cache/step4_merge_results.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Suppress warnings
warnings.filterwarnings('ignore')

@dataclass
class ReceptorData:
    """受体数据结构"""
    receptor_id: str
    uniprot_id: str
    gene_name: str
    organism: str
    pdb_id: str
    
    # Step 1 data (STRING)
    reliability_score: float
    literature_support: int = 0
    
    # Step 2 data (Docking)
    avg_binding_energy: float = 0.0
    success_rate: float = 0.0
    total_conformations: int = 0
    high_affinity: bool = False
    docking_date: str = ""
    
    # Step 3 data (Conservation)
    species_count: int = 0
    species_names: str = ""
    avg_conservation: float = 0.0
    max_identity: float = 0.0
    min_identity: float = 0.0
    is_conservative: bool = False
    alignment_length: int = 0
    conservation_threshold: float = 0.8
    
    # Calculated scores
    binding_energy_score: float = 0.0
    conservation_score: float = 0.0
    literature_score: float = 0.0
    composite_score: float = 0.0
    priority_rank: int = 0
    is_core_receptor: bool = False

class ResultsMerger:
    """结果合并和优先级评估器"""
    
    def __init__(self, config_file: str = "config/config.json"):
        """初始化合并器"""
        self.config_file = config_file
        self.config = self._load_config()
        self.cache_dir = Path(self.config['paths']['cache_dir'])
        self.output_dir = Path('./output')
        
        # Create output directory
        self.output_dir.mkdir(exist_ok=True)
        
        # Neo4j configuration
        self.neo4j_config = {
            'uri': os.getenv('NEO4J_URI', 'bolt://localhost:7687'),
            'user': os.getenv('NEO4J_USER', 'neo4j'),
            'password': os.getenv('NEO4J_PASSWORD', 'password')
        }
        
        # Weight configuration
        self.weights = {
            'binding_energy': 0.40,    # 结合能权重 40%
            'conservation': 0.30,      # 保守性权重 30%
            'literature': 0.30         # 文献支持度权重 30%
        }
        
        # Priority thresholds
        self.top_n_count = 3  # Top 3 受体标记为核心受体
        
        logger.info("ResultsMerger initialized successfully")
    
    def _load_config(self) -> Dict[str, Any]:
        """加载配置文件"""
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.error(f"Filters file not found: {self.config_file}")
            return {}
        except json.JSONDecodeError as e:
            logger.error(f"Error parsing JSON config: {e}")
            return {}
    
    def load_step_results(self) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """加载步骤1-3的结果文件"""
        logger.info("Loading step results...")
        
        # Step 1: STRING interaction results
        string_file = self.cache_dir / "string_receptors.csv"
        if not string_file.exists():
            raise FileNotFoundError(f"Step 1 results not found: {string_file}")
        string_df = pd.read_csv(string_file)
        logger.info(f"Step 1: Loaded {len(string_df)} STRING receptors")
        
        # Step 2: Docking results
        docking_file = self.cache_dir / "docking_results.csv"
        if not docking_file.exists():
            raise FileNotFoundError(f"Step 2 results not found: {docking_file}")
        docking_df = pd.read_csv(docking_file)
        logger.info(f"Step 2: Loaded {len(docking_df)} docking results")
        
        # Step 3: Conservation results
        conservation_file = self.cache_dir / "conservation_results.csv"
        if not conservation_file.exists():
            raise FileNotFoundError(f"Step 3 results not found: {conservation_file}")
        conservation_df = pd.read_csv(conservation_file)
        logger.info(f"Step 3: Loaded {len(conservation_df)} conservation results")
        
        return string_df, docking_df, conservation_df
    
    def merge_data(self, string_df: pd.DataFrame, docking_df: pd.DataFrame, 
                   conservation_df: pd.DataFrame) -> pd.DataFrame:
        """合并三个步骤的数据"""
        logger.info("Merging three-step results...")
        
        # Merge STRING and docking results on receptor_id  
        merged_df = pd.merge(
            string_df, docking_df, 
            on='receptor_id', 
            how='inner', 
            suffixes=('', '_docking')
        )
        logger.info(f"After STRING-docking merge: {len(merged_df)} receptors")
        
        # Clean up duplicate columns from STRING-docking merge
        for col in ['uniprot_id', 'gene_name', 'organism', 'pdb_id']:
            if f'{col}_docking' in merged_df.columns:
                merged_df[f'{col}'] = merged_df[f'{col}_docking']  # Keep the docking values
                merged_df = merged_df.drop(columns=[f'{col}_docking'])
        
        # Merge with conservation results on receptor_id
        # Note: Conservation results may use different ID format (gene_name vs receptor_id)
        # First try direct merge, if no matches, try merging on gene_name
        initial_merge = pd.merge(
            merged_df, conservation_df,
            left_on='receptor_id', 
            right_on='receptor_id',
            how='inner',
            suffixes=('', '_conservation')
        )
        
        if len(initial_merge) == 0:
            # Try merging on gene_name if direct merge failed
            logger.info("Direct receptor_id merge failed, trying gene_name merge...")
            
            # Create a copy of conservation_df with cleaned receptor_id for matching
            conservation_clean = conservation_df.copy()
            conservation_clean['gene_name_from_receptor_id'] = conservation_clean['receptor_id'].str.replace('_DEMO', '')
            
            merged_df = pd.merge(
                merged_df, conservation_clean,
                left_on='gene_name', 
                right_on='gene_name_from_receptor_id',
                how='inner',
                suffixes=('', '_conservation')
            )
            
            # Remove the temporary column
            if 'gene_name_from_receptor_id' in merged_df.columns:
                merged_df = merged_df.drop(columns=['gene_name_from_receptor_id'])
        else:
            merged_df = initial_merge
        logger.info(f"After conservation merge: {len(merged_df)} receptors")
        
        # Remove conservation-related duplicate columns (keep original conservation data)
        conservation_cleanup_cols = ['receptor_id_conservation', 'species_names_conservation']
        for col in conservation_cleanup_cols:
            if col in merged_df.columns:
                merged_df = merged_df.drop(columns=[col])
        
        # Ensure data types are correct
        numeric_cols = ['reliability_score', 'avg_binding_energy', 'success_rate', 
                       'avg_conservation', 'max_identity', 'min_identity']
        for col in numeric_cols:
            if col in merged_df.columns:
                merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce')
        
        # Convert boolean columns
        bool_cols = ['high_affinity', 'is_conservative']
        for col in bool_cols:
            if col in merged_df.columns:
                merged_df[col] = merged_df[col].astype(bool)
        
        logger.info(f"Final merged data: {len(merged_df)} receptors with all three analyses")
        return merged_df
    
    def calculate_priority_scores(self, merged_df: pd.DataFrame) -> pd.DataFrame:
        """计算优先级评分"""
        logger.info("Calculating priority scores...")
        
        # Calculate binding energy score (normalized, higher is better)
        # Binding energy is negative, so we normalize from 0 to 1
        binding_energy_min = merged_df['avg_binding_energy'].min()
        binding_energy_max = merged_df['avg_binding_energy'].max()
        merged_df['binding_energy_score'] = (
            (merged_df['avg_binding_energy'] - binding_energy_min) / 
            (binding_energy_max - binding_energy_min)
        )
        
        # Calculate conservation score (already normalized)
        merged_df['conservation_score'] = merged_df['avg_conservation']
        
        # Calculate literature support得分 (基于reliability_score)
        merged_df['literature_score'] = merged_df['reliability_score']
        
        # Calculate composite score
        merged_df['composite_score'] = (
            merged_df['binding_energy_score'] * self.weights['binding_energy'] +
            merged_df['conservation_score'] * self.weights['conservation'] +
            merged_df['literature_score'] * self.weights['literature']
        )
        
        # Sort by composite score (descending)
        merged_df = merged_df.sort_values('composite_score', ascending=False, ignore_index=True)
        
        # Add rank and mark core receptors
        merged_df['priority_rank'] = range(1, len(merged_df) + 1)
        merged_df['is_core_receptor'] = merged_df['priority_rank'] <= self.top_n_count
        
        logger.info(f"Priority scoring completed. Top receptor: {merged_df.iloc[0]['receptor_id']}")
        logger.info(f"Core receptors: {list(merged_df[merged_df['is_core_receptor']]['receptor_id'])}")
        
        return merged_df
    
    def create_neo4j_graph(self, merged_df: pd.DataFrame) -> Dict[str, int]:
        """创建Neo4j知识图谱"""
        logger.info("Creating Neo4j knowledge graph...")
        
        stats = {'nodes_created': 0, 'relationships_created': 0, 'errors': 0}
        
        if not NEO4J_AVAILABLE:
            logger.warning("Neo4j integration not available. Skipping knowledge graph creation.")
            return stats
        
        try:
            # Connect to Neo4j
            graph = Graph(
                uri=self.neo4j_config['uri'],
                user=self.neo4j_config['user'],
                password=self.neo4j_config['password']
            )
            
            # Get target protein info from config
            target_protein = self.config.get('target_protein', {})
            protein_name = target_protein.get('name', 'THBS4')
            
            # Clear existing graph (optional - remove if you want to keep existing data)
            # graph.delete_all()
            
            # Create protein node
            protein_node = Node("Protein", 
                              name=protein_name,
                              type="target_protein",
                              source="step4_merge")
            graph.merge(protein_node, "Protein", "name")
            
            # Create receptor nodes and relationships
            for _, row in merged_df.iterrows():
                try:
                    # Create receptor node
                    receptor_node = Node("Receptor",
                                       receptor_id=row['receptor_id'],
                                       uniprot_id=row['uniprot_id'],
                                       gene_name=row['gene_name'],
                                       organism=row['organism'],
                                       pdb_id=str(row.get('pdb_id', '')),
                                       composite_score=float(row['composite_score']),
                                       priority_rank=int(row['priority_rank']),
                                       is_core_receptor=bool(row['is_core_receptor']),
                                       source="step4_merge")
                    graph.merge(receptor_node, "Receptor", "receptor_id")
                    stats['nodes_created'] += 1
                    
                    # Create binding relationship
                    binding_rel = Relationship(protein_node, "BINDS_WITH",
                                             receptor_node, 
                                             binding_energy=float(row['avg_binding_energy']),
                                             success_rate=float(row['success_rate']),
                                             high_affinity=bool(row['high_affinity']))
                    graph.create(binding_rel)
                    stats['relationships_created'] += 1
                    
                    # Create conservation relationship (if applicable)
                    if row.get('is_conservative', False) and pd.notna(row.get('species_names')):
                        # Create homologue receptor node
                        homologue_node = Node("Receptor",
                                            receptor_id=f"{row['receptor_id']}_homologue",
                                            gene_name=row['gene_name'],
                                            organism="homologue",
                                            is_conservative=bool(row['is_conservative']),
                                            source="step4_merge")
                        graph.merge(homologue_node, "Receptor", "receptor_id")
                        stats['nodes_created'] += 1
                        
                        # Create conservation relationship
                        conservation_rel = Relationship(receptor_node, "CONSERVED_WITH",
                                                     homologue_node,
                                                     conservation_rate=float(row['avg_conservation']),
                                                     species_count=int(row.get('species_count', 0)),
                                                     species_names=str(row.get('species_names', '')))
                        graph.create(conservation_rel)
                        stats['relationships_created'] += 1
                        
                except Exception as e:
                    logger.error(f"Error creating Neo4j data for receptor {row['receptor_id']}: {e}")
                    stats['errors'] += 1
            
            logger.info(f"Neo4j operations completed: {stats}")
            
        except Exception as e:
            logger.error(f"Neo4j connection error: {e}")
            stats['errors'] += 1
            
        finally:
            if graph:
                graph.close()
        
        return stats
    
    def generate_excel_report(self, merged_df: pd.DataFrame) -> str:
        """生成Excel格式的最终报告"""
        logger.info("Generating Excel report...")
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        
        # Add summary information
        summary_data = [
            ['Receptor Priority Analysis Report'],
            [''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Receptors Analyzed:', len(merged_df)],
            ['Core Receptors (Top 3):', len(merged_df[merged_df['is_core_receptor']])],
            [''],
            ['Weights Configuration:'],
            ['Binding Energy:', f"{self.weights['binding_energy']:.1%}"],
            ['Conservation:', f"{self.weights['conservation']:.1%}"],
            ['Literature Support:', f"{self.weights['literature']:.1%}"],
            [''],
            ['Top Receptors:']
        ]
        
        # Add top receptor information
        top_receptors = merged_df.head(10)
        summary_data.append(['Rank', 'Receptor ID', 'Gene Name', 'Composite Score', 'Status'])
        for _, row in top_receptors.iterrows():
            status = 'Core Receptor' if row['is_core_receptor'] else 'Secondary'
            summary_data.append([
                row['priority_rank'],
                row['receptor_id'],
                row['gene_name'],
                f"{row['composite_score']:.3f}",
                status
            ])
        
        # Write summary data
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                
                # Style headers
                if row_num == 1:  # Title
                    cell.font = Font(bold=True, size=16)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=16, color="FFFFFF")
                elif row_num in [2, 7, 13]:  # Sub-headers
                    cell.font = Font(bold=True)
                elif row_num == 13 and col_num == 1:  # Column headers
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    cell.font = Font(bold=True)
        
        # Create detailed data sheet
        details_sheet = workbook.create_sheet("Detailed Results")
        
        # Prepare data for DataFrame conversion
        output_df = merged_df[[
            'priority_rank', 'receptor_id', 'uniprot_id', 'gene_name', 'organism',
            'reliability_score', 'avg_binding_energy', 'success_rate', 'total_conformations', 'high_affinity',
            'avg_conservation', 'max_identity', 'min_identity', 'is_conservative', 'species_count',
            'binding_energy_score', 'conservation_score', 'literature_score', 'composite_score',
            'is_core_receptor'
        ]].copy()
        
        # Rename columns for better readability
        column_mapping = {
            'priority_rank': '排名',
            'receptor_id': '受体ID',
            'uniprot_id': 'UniProt ID',
            'gene_name': '基因名称',
            'organism': '物种',
            'reliability_score': '可靠性评分',
            'avg_binding_energy': '平均结合能',
            'success_rate': '成功率',
            'total_conformations': '总构象数',
            'high_affinity': '高亲和性',
            'avg_conservation': '保守性评分',
            'max_identity': '最大一致性',
            'min_identity': '最小一致性',
            'is_conservative': '为保守蛋白',
            'species_count': '物种数量',
            'binding_energy_score': '结合能得分',
            'conservation_score': '保守性得分',
            'literature_score': '文献得分',
            'composite_score': '综合得分',
            'is_core_receptor': '核心受体'
        }
        
        output_df = output_df.rename(columns=column_mapping)
        
        # Write DataFrame to Excel
        for row in dataframe_to_rows(output_df, index=False, header=True):
            details_sheet.append(row)
        
        # Style the detailed sheet
        self._style_excel_sheet(details_sheet, len(output_df.columns))
        
        # Save workbook
        report_file = self.output_dir / "receptor_priority.xlsx"
        workbook.save(report_file)
        
        logger.info(f"Excel report saved: {report_file}")
        return str(report_file)
    
    def _style_excel_sheet(self, sheet, num_cols):
        """样式化Excel工作表"""
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Style header row
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style data rows
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def run_analysis(self) -> Dict[str, Any]:
        """运行完整的合并分析"""
        logger.info("Starting Step 4: Results Merging and Priority Ranking...")
        
        results = {
            'status': 'success',
            'start_time': datetime.now().isoformat(),
            'step1_count': 0,
            'step2_count': 0,
            'step3_count': 0,
            'merged_count': 0,
            'core_receptors': [],
            'neo4j_stats': {},
            'excel_report': '',
            'errors': []
        }
        
        try:
            # Load data from all three steps
            string_df, docking_df, conservation_df = self.load_step_results()
            results['step1_count'] = len(string_df)
            results['step2_count'] = len(docking_df)
            results['step3_count'] = len(conservation_df)
            
            # Merge data
            merged_df = self.merge_data(string_df, docking_df, conservation_df)
            results['merged_count'] = len(merged_df)
            
            # Calculate priority scores
            merged_df = self.calculate_priority_scores(merged_df)
            
            # Identify core receptors
            core_receptors = merged_df[merged_df['is_core_receptor']][
                ['priority_rank', 'receptor_id', 'gene_name', 'composite_score']
            ].to_dict('records')
            results['core_receptors'] = core_receptors
            
            # Create Neo4j knowledge graph
            neo4j_stats = self.create_neo4j_graph(merged_df)
            results['neo4j_stats'] = neo4j_stats
            
            # Generate Excel report
            excel_file = self.generate_excel_report(merged_df)
            results['excel_report'] = excel_file
            
            logger.info(f"Step 4 completed successfully:")
            logger.info(f"- Merged {results['merged_count']} receptors from all three steps")
            logger.info(f"- Identified {len(core_receptors)} core receptors")
            logger.info(f"- Created Neo4j graph with {neo4j_stats['relationships_created']} relationships")
            logger.info(f"- Generated Excel report: {excel_file}")
            
        except Exception as e:
            logger.error(f"Step 4 analysis failed: {e}")
            results['status'] = 'failed'
            results['errors'].append(str(e))
        
        finally:
            results['end_time'] = datetime.now().isoformat()
        
        return results

def main():
    """主函数"""
    try:
        # Initialize merger
        merger = ResultsMerger()
        
        # Run analysis
        results = merger.run_analysis()
        
        # Print summary
        print("\n" + "="*60)
        print("STEP 4: RESULTS MERGING AND PRIORITY RANKING")
        print("="*60)
        
        if results['status'] == 'success':
            print(f"✓ Analysis completed successfully!")
            print(f"✓ Total receptors analyzed: {results['merged_count']}")
            print(f"✓ Core receptors identified: {len(results['core_receptors'])}")
            print(f"✓ Neo4j relationships created: {results['neo4j_stats'].get('relationships_created', 0)}")
            print(f"✓ Excel report generated: {results['excel_report']}")
            
            print(f"\nCore Receptors (Top {merger.top_n_count}):")
            for receptor in results['core_receptors']:
                print(f"  {receptor['priority_rank']}. {receptor['receptor_id']} ({receptor['gene_name']}) "
                      f"- Score: {receptor['composite_score']:.3f}")
        else:
            print(f"✗ Analysis failed: {results['errors']}")
            
    except KeyboardInterrupt:
        print("\nAnalysis interrupted by user.")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(f"✗ Unexpected error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
